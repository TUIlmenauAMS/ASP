# -*- coding: utf-8 -*-
__author__ = 'S.I. Mimilakis'
__copyright__ = 'MacSeNet'

import IOMethods as IO
import TFMethods as TF
from MaskingMethods import FrequencyMasking as fm
import numpy as np
import scipy.signal as sig
import os
import QMF.qmf_realtime_class as qrf
from multiprocessing import Process
import matplotlib.pyplot as plt

eps = np.finfo(np.double).tiny

# STFT
def eval_stft_hann(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.hanning(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.hanning(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_bt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.bartlett(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.bartlett(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_nt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_ntB(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_stft_hann(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.hanning(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.hanning(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_bt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.bartlett(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.bartlett(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_nt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_ntB(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

# MDCT
def eval_mdct(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    MixmX = np.real(MixmX)
    VoxmX = np.real(VoxmX)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_mdcst_complex(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    # Acquire Magnitude from Complex
    MixmX = np.abs(MixmX)
    VoxmX = np.abs(VoxmX)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_mdct(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    MixmX = np.real(MixmX)
    VoxmX = np.real(VoxmX)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_mdcst_complex(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    # Acquire Magnitude from Complex
    MixmX = np.abs(MixmX)
    VoxmX = np.abs(VoxmX)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

# PQMF
def eval_pqmf_cos(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.float32)
    vs = np.empty((timeSlots, N), dtype = np.float32)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(karms + vs), np.abs(vs), np.abs(karms), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(karms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_pqmf_complex(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.complex64)
    vs = np.empty((timeSlots, N), dtype = np.complex64)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(karms + vs), np.abs(vs), np.abs(karms), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(karms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_pqmf_cos(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.float32)
    vs = np.empty((timeSlots, N), dtype = np.float32)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(karms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

def spc_eval_pqmf_complex(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.complex64)
    vs = np.empty((timeSlots, N), dtype = np.complex64)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(karms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

# Result Loading
def load_results():
    # STFT
    # Hanning
    stft_hann_a1 = np.load('WDO-Experiment/stft_hann_a1.npy')
    stft_hann_a12 = np.load('WDO-Experiment/stft_hann_a12.npy')

    # Bartlett
    stft_bt_a1 = np.load('WDO-Experiment/stft_bt_a1.npy')
    stft_bt_a12 = np.load('WDO-Experiment/stft_bt_a12.npy')

    # Nuttall-4b
    stft_nt_a1 = np.load('WDO-Experiment/stft_nt_a1.npy')
    stft_nt_a12 = np.load('WDO-Experiment/stft_nt_a12.npy')

    stft_ntB_a1 = np.load('WDO-Experiment/stft_ntB_a1.npy')
    stft_ntB_a12 = np.load('WDO-Experiment/stft_ntB_a12.npy')


    # MDCT
    mdct_a1 = np.load('WDO-Experiment/mdct_a1.npy')
    mdct_a12 = np.load('WDO-Experiment/mdct_a12.npy')
    mdcst_a1 = np.load('WDO-Experiment/mdcst_a1.npy')
    mdcst_a12 = np.load('WDO-Experiment/mdcst_a12.npy')
    # PQMF
    pqmf_cos_a1 = np.load('WDO-Experiment/pqmf_cos_a1.npy')
    pqmf_cos_a2 = np.load('WDO-Experiment/pqmf_cos_a12.npy')
    pqmf_compl_a1 = np.load('WDO-Experiment/pqmf_compl_a1.npy')
    pqmf_compl_a2 = np.load('WDO-Experiment/pqmf_compl_a12.npy')

    return stft_hann_a1,stft_hann_a12,stft_bt_a1, stft_bt_a12, stft_nt_a1,stft_nt_a12, stft_ntB_a1, stft_ntB_a12, \
    mdct_a1, mdct_a12, mdcst_a1, mdcst_a12, pqmf_cos_a1, pqmf_cos_a2, pqmf_compl_a1, pqmf_compl_a2

def load_SPCresults():
    # STFT
    # Hanning
    spc_stft_hann = np.load('WDO-Experiment/spc_stft_hann_a1.npy')

    # Bartlett
    spc_stft_bt = np.load('WDO-Experiment/spc_stft_bt_a1.npy')

    # Nuttall-4b
    spc_stft_nt = np.load('WDO-Experiment/spc_stft_nt_a1.npy')
    spc_stft_ntB = np.load('WDO-Experiment/spc_stft_ntB_a1.npy')

    # MDCT
    spc_mdct = np.load('WDO-Experiment/spc_mdct_a1.npy')
    spc_mdcst = np.load('WDO-Experiment/spc_mdcst_a1.npy')

    # PQMF
    spc_pqmf = np.load('WDO-Experiment/spc_pqmf_cos_a1.npy')
    spc_pqmf_comp = np.load('WDO-Experiment/spc_pqmf_compl_a1.npy')

    return spc_stft_hann, spc_stft_bt, spc_stft_nt, spc_stft_ntB, spc_mdct,\
           spc_mdcst, spc_pqmf, spc_pqmf_comp

# Main Operations
### WDO
def mainWDO(selection):
    from openpyxl import load_workbook
    MixturesPath = '/home/avdata/audio/own/dsd100/DSD100/Mixtures/'
    SourcesPath = '/home/avdata/audio/own/dsd100/DSD100/Sources/'
    savePath = '/mnt/IDMT-WORKSPACE/DATA-STORE/mis/Datasets/DSD/mono_sv_eval'
    foldersList = ['Dev', 'Test']

    xlsLoc = '/home/avdata/audio/own/dsd100/dsd100.xlsx'
    keywords = ['bass_seg.wav', 'drums_seg.wav', 'other_seg.wav', 'vocals_seg.wav']
    wb = load_workbook(filename = xlsLoc)
    sheet_ranges = wb['Sheet1']

    # Worksheet Indices
    strtIndx = 2
    endIndx = 102

    # Saving Index for storing the results
    saveIndx = 0

    # Number of sub-bands
    N = 1024

    # Initialize Results storing
    # Alpha = 1.
    stft_hann_a1 = np.zeros(endIndx - strtIndx)
    stft_bt_a1 = np.zeros(endIndx - strtIndx)
    stft_nt_a1 = np.zeros(endIndx - strtIndx)
    stft_ntB_a1 = np.zeros(endIndx - strtIndx)

    mdct_a1 = np.zeros(endIndx - strtIndx)
    mdcst_a1 = np.zeros(endIndx - strtIndx)
    pqmf_cos_a1 = np.zeros(endIndx - strtIndx)
    pqmf_compl_a1 = np.zeros(endIndx - strtIndx)

    # Alpha = 1.2
    stft_hann_a12 = np.zeros(endIndx - strtIndx)
    stft_bt_a12 = np.zeros(endIndx - strtIndx)
    stft_nt_a12 = np.zeros(endIndx - strtIndx)
    stft_ntB_a12 = np.zeros(endIndx - strtIndx)

    mdct_a12 = np.zeros(endIndx - strtIndx)
    mdcst_a12 = np.zeros(endIndx - strtIndx)
    pqmf_cos_a12 = np.zeros(endIndx - strtIndx)
    pqmf_compl_a12 = np.zeros(endIndx - strtIndx)

    for Fileindx in range(strtIndx,endIndx):
        cell = 'A' + str(Fileindx)
        print(cell)
        strr =(sheet_ranges[cell].value)
        strrb = (sheet_ranges[cell].value)
        DevfolderName = os.path.join(MixturesPath,foldersList[0],str(strr.encode('utf-8')))
        TstfolderName = os.path.join(MixturesPath,foldersList[1],str(strrb.encode('utf-8')))

        if os.path.exists(TstfolderName):
            print('Current File belongs to Test/Validation Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[1],str(sheet_ranges[cell].value))
            folderNameMix = TstfolderName
            currFile = 'Test'

        elif os.path.exists(DevfolderName):
            print('Current File belongs to Training Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[0],str(sheet_ranges[cell].value))
            folderNameMix = DevfolderName
            currFile = 'Valid'

        print('Reading')
        mix, fs = IO.AudioIO.wavRead(os.path.join(folderNameMix, 'mixture_seg.wav'), mono = True)
        bss, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[0]), mono = True)
        drm, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[1]), mono = True)
        oth, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[2]), mono = True)
        mixKaraoke = bss + drm + oth
        vox, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[3]), mono = True)

        # Evaluation inside file loop
        if selection == 'stft':
            print('STFT')
            # Alpha = 1.
            stft_hann_a1[saveIndx] = eval_stft_hann(1., mixKaraoke, vox)
            stft_bt_a1[saveIndx] = eval_stft_bt(1., mixKaraoke, vox)
            stft_nt_a1[saveIndx] = eval_stft_nt(1., mixKaraoke, vox)
            stft_ntB_a1[saveIndx] = eval_stft_ntB(1., mixKaraoke, vox)


            # Alpha = 1.2
            stft_hann_a12[saveIndx] = eval_stft_hann(1.2, mixKaraoke, vox)
            stft_bt_a12[saveIndx] = eval_stft_bt(1.2, mixKaraoke, vox)
            stft_nt_a12[saveIndx] = eval_stft_nt(1.2, mixKaraoke, vox)
            stft_nt_a12[saveIndx] = eval_stft_ntB(1.2, mixKaraoke, vox)

            np.save('WDO-Experiment/stft_hann_a1.npy', stft_hann_a1)
            np.save('WDO-Experiment/stft_bt_a1.npy', stft_bt_a1)
            np.save('WDO-Experiment/stft_nt_a1.npy', stft_nt_a1)
            np.save('WDO-Experiment/stft_ntB_a1.npy', stft_ntB_a1)
            np.save('WDO-Experiment/stft_hann_a12.npy', stft_hann_a12)
            np.save('WDO-Experiment/stft_bt_a12.npy', stft_bt_a12)
            np.save('WDO-Experiment/stft_nt_a12.npy', stft_nt_a12)
            np.save('WDO-Experiment/stft_ntB_a12.npy', stft_ntB_a12)


        elif selection == 'mdct':
            print('MDCT')
            # Alpha = 1.
            mdct_a1[saveIndx] = eval_mdct(1., mixKaraoke, vox)
            mdcst_a1[saveIndx] = eval_mdcst_complex(1., mixKaraoke, vox)

            # Alpha = 1.2
            mdct_a12[saveIndx] = eval_mdct(1.2, mixKaraoke, vox)
            mdcst_a12[saveIndx] = eval_mdcst_complex(1.2, mixKaraoke, vox)

            np.save('WDO-Experiment/mdct_a1.npy', mdct_a1)
            np.save('WDO-Experiment/mdcst_a1.npy', mdcst_a1)
            np.save('WDO-Experiment/mdct_a12.npy', mdct_a12)
            np.save('WDO-Experiment/mdcst_a12.npy', mdcst_a12)

        elif selection == 'pqmf':
            print('PQMF')
            # Alpha = 1.
            pqmf_cos_a1[saveIndx] = eval_pqmf_cos(1., mixKaraoke, vox)

            # Alpha = 1.2
            pqmf_cos_a12[saveIndx] = eval_pqmf_cos(1.2, mixKaraoke, vox)

            np.save('WDO-Experiment/pqmf_cos_a1.npy', pqmf_cos_a1)
            np.save('WDO-Experiment/pqmf_cos_a12.npy', pqmf_cos_a12)


        elif selection == 'pqmf_complex':
            print('PQMF Complex')
            # Alpha = 1.
            pqmf_compl_a1[saveIndx] = eval_pqmf_complex(1., mixKaraoke, vox)

            # Alpha = 1.2
            pqmf_compl_a12[saveIndx] = eval_pqmf_complex(1.2, mixKaraoke, vox)

            np.save('WDO-Experiment/pqmf_compl_a1.npy', pqmf_compl_a1)
            np.save('WDO-Experiment/pqmf_compl_a12.npy', pqmf_compl_a12)

        else :
            assert('Unknown Selection!')

        # Update Storing Index
        saveIndx += 1

    return None

### SPC
def mainSPC(selection):
    from openpyxl import load_workbook
    MixturesPath = '/home/avdata/audio/own/dsd100/DSD100/Mixtures/'
    SourcesPath = '/home/avdata/audio/own/dsd100/DSD100/Sources/'
    savePath = '/mnt/IDMT-WORKSPACE/DATA-STORE/mis/Datasets/DSD/mono_sv_eval'
    foldersList = ['Dev', 'Test']

    xlsLoc = '/home/avdata/audio/own/dsd100/dsd100.xlsx'
    keywords = ['bass_seg.wav', 'drums_seg.wav', 'other_seg.wav', 'vocals_seg.wav']
    wb = load_workbook(filename = xlsLoc)
    sheet_ranges = wb['Sheet1']

    # Worksheet Indices
    strtIndx = 2
    endIndx = 102

    # Saving Index for storing the results
    saveIndx = 0

    # Number of sub-bands
    N = 1024

    # Initialize Results storing
    # Alpha = 1.
    stft_hann_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_bt_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_nt_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_ntB_a1 = np.zeros((endIndx - strtIndx, 2))

    mdct_a1 = np.zeros((endIndx - strtIndx, 2))
    mdcst_a1 = np.zeros((endIndx - strtIndx, 2))
    pqmf_cos_a1 = np.zeros((endIndx - strtIndx, 2))
    pqmf_compl_a1 = np.zeros((endIndx - strtIndx, 2))

    for Fileindx in range(strtIndx,endIndx):
        cell = 'A' + str(Fileindx)
        print(cell)
        strr =(sheet_ranges[cell].value)
        strrb = (sheet_ranges[cell].value)
        DevfolderName = os.path.join(MixturesPath,foldersList[0],str(strr.encode('utf-8')))
        TstfolderName = os.path.join(MixturesPath,foldersList[1],str(strrb.encode('utf-8')))

        if os.path.exists(TstfolderName):
            print('Current File belongs to Test/Validation Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[1],str(sheet_ranges[cell].value))
            folderNameMix = TstfolderName
            currFile = 'Test'

        elif os.path.exists(DevfolderName):
            print('Current File belongs to Training Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[0],str(sheet_ranges[cell].value))
            folderNameMix = DevfolderName
            currFile = 'Valid'

        print('Reading')
        mix, fs = IO.AudioIO.wavRead(os.path.join(folderNameMix, 'mixture_seg.wav'), mono = True)
        bss, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[0]), mono = True)
        drm, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[1]), mono = True)
        oth, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[2]), mono = True)
        mixKaraoke = bss + drm + oth
        vox, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[3]), mono = True)

        # Evaluation inside file loop
        if selection == 'stft':
            print('STFT')
            # Alpha = 1.
            stft_hann_a1[saveIndx, 0], stft_hann_a1[saveIndx, 1] = spc_eval_stft_hann(1., mixKaraoke, vox)
            stft_bt_a1[saveIndx, 0], stft_bt_a1[saveIndx, 1] = spc_eval_stft_bt(1., mixKaraoke, vox)
            stft_nt_a1[saveIndx, 0], stft_nt_a1[saveIndx, 1] = spc_eval_stft_nt(1., mixKaraoke, vox)
            stft_ntB_a1[saveIndx, 0], stft_ntB_a1[saveIndx, 1] = spc_eval_stft_ntB(1., mixKaraoke, vox)

            np.save('WDO-Experiment/spc_stft_hann_a1.npy', stft_hann_a1)
            np.save('WDO-Experiment/spc_stft_bt_a1.npy', stft_bt_a1)
            np.save('WDO-Experiment/spc_stft_nt_a1.npy', stft_nt_a1)
            np.save('WDO-Experiment/spc_stft_ntB_a1.npy', stft_ntB_a1)


        elif selection == 'mdct':
            print('MDCT')
            # Alpha = 1.
            mdct_a1[saveIndx, 0], mdct_a1[saveIndx, 1] = spc_eval_mdct(1., mixKaraoke, vox)
            mdcst_a1[saveIndx, 0], mdcst_a1[saveIndx, 1] = spc_eval_mdcst_complex(1., mixKaraoke, vox)

            np.save('WDO-Experiment/spc_mdct_a1.npy', mdct_a1)
            np.save('WDO-Experiment/spc_mdcst_a1.npy', mdcst_a1)


        elif selection == 'pqmf':
            print('PQMF')
            # Alpha = 1.
            pqmf_cos_a1[saveIndx, 0], pqmf_cos_a1[saveIndx, 1] = spc_eval_pqmf_cos(1., mixKaraoke, vox)

            np.save('WDO-Experiment/spc_pqmf_cos_a1.npy', pqmf_cos_a1)
            np.save('WDO-Experiment/spc_pqmf_cos_a12.npy', pqmf_cos_a12)


        elif selection == 'pqmf_complex':
            print('PQMF Complex')
            # Alpha = 1.
            pqmf_compl_a1[saveIndx, 0], pqmf_compl_a1[saveIndx, 1] = spc_eval_pqmf_complex(1., mixKaraoke, vox)

            np.save('WDO-Experiment/spc_pqmf_compl_a1.npy', pqmf_compl_a1)

        else :
            assert('Unknown Selection!')

        # Update Storing Index
        saveIndx += 1

    return None

if __name__ == "__main__":
    # Define Operation
    operation = 'Results'
    multi_processing = True

    if operation == 'WDO' :
        if multi_processing == True :
            p1 = Process(target = mainWDO, args = ('stft',))
            p1.start()
            p2 = Process(target = mainWDO, args = ('mdct',))
            p2.start()
            p3 = Process(target = mainWDO, args = ('pqmf',))
            p3.start()
            p4 = Process(target = mainWDO, args = ('pqmf_complex',))
            p4.start()

            p1.join()
            p2.join()
            p3.join()
            p4.join()
        else :
            mainWDO()

    if operation == 'Sparsity' :
        if multi_processing == True :
            p1 = Process(target = mainSPC, args = ('stft',))
            p1.start()
            p2 = Process(target = mainSPC, args = ('mdct',))
            p2.start()
            p3 = Process(target = mainSPC, args = ('pqmf',))
            p3.start()
            p4 = Process(target = mainSPC, args = ('pqmf_complex',))
            p4.start()

            p1.join()
            p2.join()
            p3.join()
            p4.join()
        else :
            mainSPC()

    if operation == 'Results':
        print('Loading WDO Results')

        stft_hann_a1, stft_hann_a12, stft_bt_a1, stft_bt_a12, stft_nt_a1, stft_nt_a12, stft_ntB_a1, stft_ntB_a12, \
        mdct_a1, mdct_a12, mdcst_a1, mdcst_a12, pqmf_cos_a1, pqmf_cos_a2, pqmf_compl_a1, pqmf_compl_a2 = load_results()

        # Avoid two audio files with silence in vocals
        stft_hann_a1 = np.delete(stft_hann_a1, (71, 75))
        stft_hann_a12 = np.delete(stft_hann_a12, (71, 75))
        stft_bt_a1 = np.delete(stft_bt_a1, (71, 75))
        stft_bt_a12 = np.delete(stft_bt_a12, (71, 75))
        stft_nt_a1 = np.delete(stft_nt_a1, (71, 75))
        stft_nt_a12 = np.delete(stft_nt_a12, (71, 75))
        stft_ntB_a1 = np.delete(stft_ntB_a1, (71, 75))
        stft_ntB_a12 = np.delete(stft_ntB_a12, (71, 75))
        mdct_a1 = np.delete(mdct_a1, (71, 75))
        mdct_a12 = np.delete(mdct_a12, (71, 75))
        mdcst_a1 = np.delete(mdcst_a1, (71, 75))
        mdcst_a12 = np.delete(mdcst_a12, (71, 75))
        pqmf_cos_a1 = np.delete(pqmf_cos_a1, (71, 75))
        pqmf_cos_a2 = np.delete(pqmf_cos_a2, (71, 75))
        pqmf_compl_a1 = np.delete(pqmf_compl_a1, (71, 75))
        pqmf_compl_a2 = np.delete(pqmf_compl_a2, (71, 75))

        data = np.vstack((stft_hann_a1, stft_bt_a1, stft_nt_a1, stft_ntB_a1, mdct_a1, mdcst_a1, pqmf_cos_a1, pqmf_compl_a1)).T
        colors = ['cyan', 'cyan', 'lightblue', 'lightblue', 'lightgreen', 'lightgreen', 'pink', 'pink']
        labels = ['STFT Hann 50%', 'STFT Brt 50%', 'STFT Nt 75%', 'STFT Nt 50%', 'MDCT', 'MDCST', 'PQMF-cos', 'PQMF-complex']

        meanpointprops = dict(marker='D', markeredgecolor='black', markerfacecolor='firebrick')

        # Configure Boxplot
        plt.figure(1)
        box = plt.boxplot(data, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('W-DO Measure')
        plt.show(block = False)

        print('Loading Sparsity Results')
        spc_stft_hann, spc_stft_bt, spc_stft_nt, spc_stft_ntB, spc_mdct, spc_mdcst, spc_pqmf, spc_pqmf_comp = load_SPCresults()

        # Acquire sparsity measures for mixture and singing voice
        spc_stft_hann_mix = np.delete(spc_stft_hann[:, 0], (71, 75))
        spc_stft_hann_sv = np.delete(spc_stft_hann[:, 1], (71,75))
        spc_stft_bt_mix = np.delete(spc_stft_bt[:, 0], (71, 75))
        spc_stft_bt_sv = np.delete(spc_stft_bt[:, 1], (71, 75))
        spc_stft_nt_mix = np.delete(spc_stft_nt[:, 0], (71, 75))
        spc_stft_nt_sv = np.delete(spc_stft_nt[:, 1], (71, 75))
        spc_stft_ntB_mix = np.delete(spc_stft_ntB[:, 0], (71, 75))
        spc_stft_ntB_sv = np.delete(spc_stft_ntB[:, 1], (71, 75))
        spc_mdct_mix = np.delete(spc_mdct[:, 0], (71, 75))
        spc_mdct_sv = np.delete(spc_mdct[:, 1], (71, 75))
        spc_mdcst_mix = np.delete(spc_mdcst[:, 0], (71, 75))
        spc_mdcst_sv = np.delete(spc_mdcst[:, 1], (71, 75))
        spc_pqmf_mix = np.delete(spc_pqmf[:, 0], (71, 75))
        spc_pqmf_sv = np.delete(spc_pqmf[:, 1], (71, 75))
        spc_pqmf_comp_mix = np.delete(spc_pqmf_comp[:, 0], (71, 75))
        spc_pqmf_comp_sv = np.delete(spc_pqmf_comp[:, 1], (71, 75))


        dataMIX = np.vstack((spc_stft_hann_mix, spc_stft_bt_mix, spc_stft_nt_mix, spc_stft_ntB_mix, spc_mdct_mix, spc_mdcst_mix, spc_pqmf_mix, spc_pqmf_comp_mix)).T
        dataSV = np.vstack((spc_stft_hann_sv, spc_stft_bt_sv, spc_stft_nt_sv, spc_stft_ntB_sv, spc_mdct_sv, spc_mdcst_sv, spc_pqmf_sv, spc_pqmf_comp_sv)).T

        colors = ['cyan', 'cyan', 'lightblue', 'lightblue', 'lightgreen', 'lightgreen', 'pink', 'pink']
        labels = ['STFT Hann 50%', 'STFT Brt 50%', 'STFT Nt 75%', 'STFT Nt 50%', 'MDCT', 'MDCST', 'PQMF-cos', 'PQMF-complex']

        meanpointprops = dict(marker='D', markeredgecolor='black', markerfacecolor='firebrick')

        # Configure Boxplot
        plt.figure(2)
        box2 = plt.boxplot(dataMIX, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box2['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('Sparsity Measure')
        plt.show(block = False)

        # Configure Boxplot
        plt.figure(3)
        box3 = plt.boxplot(dataSV, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box3['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('Sparsity Measure')
        plt.show(block = False)






