# -*- coding: utf-8 -*-
__author__ = 'S.I. Mimilakis'
__copyright__ = 'MacSeNet'

import IOMethods as IO
import TFMethods as TF
from MaskingMethods import FrequencyMasking as fm
import numpy as np
import scipy.signal as sig
import os
import QMF.qmf_realtime_class as qrf
from multiprocessing import Process
import matplotlib.pyplot as plt
import spams, math

eps = np.finfo(np.double).tiny

# STFT
def eval_stft_hann(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.hanning(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.hanning(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_bt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.bartlett(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.bartlett(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_nt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_stft_ntB(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_stft_hann(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.hanning(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.hanning(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_bt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = sig.bartlett(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = sig.bartlett(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_nt(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 256)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_stft_ntB(alpha, mixKaraoke, vox):
    MixmX, _ = TF.TimeFrequencyDecomposition.STFT(mixKaraoke, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)
    VoxmX, _ = TF.TimeFrequencyDecomposition.STFT(vox, w = TF.TimeFrequencyDecomposition.nuttall4b(1024, False), N = 1024, hop = 512)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

# MDCT
def eval_mdct(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    MixmX = np.real(MixmX)
    VoxmX = np.real(VoxmX)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_mdcst_complex(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    # Acquire Magnitude from Complex
    MixmX = np.abs(MixmX)
    VoxmX = np.abs(VoxmX)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(MixmX + VoxmX), np.abs(VoxmX), np.abs(MixmX), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(VoxmX))
    SIR = TF.WDODisjointness.SIR(M, np.abs(VoxmX), np.abs(MixmX))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_mdct(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    MixmX = np.real(MixmX)
    VoxmX = np.real(VoxmX)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

def spc_eval_mdcst_complex(alpha, mixKaraoke, vox):

    MixmX = TF.TimeFrequencyDecomposition.complex_analysis(mixKaraoke, N = 1024)
    VoxmX = TF.TimeFrequencyDecomposition.complex_analysis(vox, N = 1024)

    # Acquire Magnitude from Complex
    MixmX = np.abs(MixmX)
    VoxmX = np.abs(VoxmX)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(MixmX))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(VoxmX))

    return SPCMix, SPCVox

# PQMF
def eval_pqmf_cos(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.float32)
    vs = np.empty((timeSlots, N), dtype = np.float32)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(karms + vs), np.abs(vs), np.abs(karms), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(karms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_pqmf_complex(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.complex64)
    vs = np.empty((timeSlots, N), dtype = np.complex64)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(karms + vs), np.abs(vs), np.abs(karms), [], [], alpha = alpha, method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(karms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_pqmf_cos(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.float32)
    vs = np.empty((timeSlots, N), dtype = np.float32)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.analysisqmf_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(karms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

def spc_eval_pqmf_complex(alpha, mixKaraoke, vox):

    N = 1024
    timeSlots = len(mixKaraoke)/N
    karms = np.empty((timeSlots, N), dtype = np.complex64)
    vs = np.empty((timeSlots, N), dtype = np.complex64)

    for indx in xrange(timeSlots):
            karms[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(mixKaraoke[indx * N : (indx + 1) * N], N)

    qrf.reset_rt()
    for indx in xrange(timeSlots):
            vs[indx, :] = qrf.PQMFAnalysis.complex_analysis_realtime(vox[indx * N : (indx + 1) * N], N)

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(karms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

# OMP
def OLAnalysis(x, w, N, hop):

    # Analysis Parameters
    wsz = w.size
    print(wsz)
    hw1 = int(math.floor((wsz+1)/2))
    hw2 = int(math.floor(wsz/2))

    # Add some zeros at the start and end of the signal to avoid window smearing
    x = np.append(np.zeros(3*hop),x)
    x = np.append(x, np.zeros(3*hop))

    xol = np.empty((len(x)/hop, N))

    # Initialize sound pointers
    pin = 0

    pend = x.size - wsz

    indx = 0

    while pin <= pend:
        xSeg = x[pin:pin+wsz] * w
        xol[indx, :] = xSeg

        indx += 1
        pin += hop

    return xol
def OLAPSynth(xmX, wsz, hop):

    # Acquire half window sizes
    hw1 = int(math.floor((wsz+1)/2))
    hw2 = int(math.floor(wsz/2))

    # Acquire the number of STFT frames
    numFr = xmX.shape[0]

    # Initialise output array with zeros
    y = np.zeros(numFr * hop + hw1 + hw2)

    # Initialise sound pointer
    pin = 0

    # Main Synthesis Loop
    for indx in range(numFr):
        # Inverse Discrete Fourier Transform
        ybuffer = xmX[indx, :]

        # Overlap and Add
        y[pin:pin+wsz] += ybuffer

        # Advance pointer
        pin += hop

    # Delete the extra zeros that the analysis had placed
    y = np.delete(y, range(3*hop))
    y = np.delete(y, range(y.size-(3*hop + 1), y.size))

    return y

def eval_dct_oc(mixKaraoke, vox):
    # Analysis Parameters
    N = 1024
    winsamples = 2 * N
    win = np.sin(np.pi/(winsamples)*(np.arange(0,winsamples)+0.5))

    # Acquire modulation matrices
    Cos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N)
    Cos = Cos.T

    # Create Dictionary
    CS2 = np.asfortranarray(Cos, (np.float32))

    # Ovelapped analysis with 50%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(ms + vs), np.abs(vs), np.abs(ms), [], [], alpha = 1., method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(ms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_dctdst_oc(mixKaraoke, vox):
    # Analysis Parameters
    N = 512
    winsamples = 2 * N
    win = np.sin(np.pi/(winsamples)*(np.arange(0,winsamples)+0.5))

    # Acquire modulation matrices
    Cos, Sin = TF.TimeFrequencyDecomposition.coreModulation(win, N)
    Cos = Cos.T
    Sin = Sin.T

    # Create Dictionary
    CS2 = np.asfortranarray(np.hstack((Cos, Sin)), (np.float32))

    # Ovelapped analysis with 50%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(ms + vs), np.abs(vs), np.abs(ms), [], [], alpha = 1., method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(ms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def eval_dct_union_oc(mixKaraoke, vox):
    # Dictionary Parameters
    N = 2 ** np.arange(6,10)
    N = np.hstack((64,N))
    winsamples = 2 * N

    # Compute union of various DCT lengths for Dictionary
    for indx in xrange(N.size) :
        win = np.sin(np.pi/(winsamples[indx])*(np.arange(0, winsamples[indx]) + 0.5))
        if indx == 0:
            cCos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N[indx])
            cCos = cCos.T
            cbCos = np.zeros((2048, cCos.shape[1]))
            cbCos[:cCos.shape[0], :] = cCos
            Cos = cbCos
        else :
            cCos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N[indx])
            cCos = cCos.T
            cbCos = np.zeros((2048, cCos.shape[1]))
            cbCos[:cCos.shape[0], :] = cCos
            Cos = np.hstack((Cos, cbCos))

    # Dictionary in format for SPAMS
    CS2 = np.asfortranarray(Cos, (np.float32))

    # Parameters for the overlap analysis/synthesis
    N = 1024
    winsamples = 2048
    win = np.sin(np.pi/(winsamples)*(np.arange(0, winsamples) + 0.5))

    # Ovelapped analysis with 75%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N/4).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N/4).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Upper Bound Binary Mask
    mask = fm(np.abs(ms + vs), np.abs(vs), np.abs(ms), [], [], alpha = 1., method = 'UBBM')

    # Activate the method to acquire the mask
    vsf = mask()
    M = mask._mask

    # Compute the measures used in WDO
    PSR = TF.WDODisjointness.PSR(M, np.abs(vs))
    SIR = TF.WDODisjointness.SIR(M, np.abs(vs), np.abs(ms))

    WDO = TF.WDODisjointness.WDO(PSR, SIR)

    return WDO

def spc_eval_dct_union_oc(mixKaraoke, vox):
    # Dictionary Parameters
    N = 2 ** np.arange(6,10)
    N = np.hstack((64,N))
    winsamples = 2 * N

    # Compute union of various DCT lengths for Dictionary
    for indx in xrange(N.size) :
        win = np.sin(np.pi/(winsamples[indx])*(np.arange(0, winsamples[indx]) + 0.5))
        if indx == 0:
            cCos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N[indx])
            cCos = cCos.T
            cbCos = np.zeros((2048, cCos.shape[1]))
            cbCos[:cCos.shape[0], :] = cCos
            Cos = cbCos
        else :
            cCos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N[indx])
            cCos = cCos.T
            cbCos = np.zeros((2048, cCos.shape[1]))
            cbCos[:cCos.shape[0], :] = cCos
            Cos = np.hstack((Cos, cbCos))

    # Dictionary in format for SPAMS
    CS2 = np.asfortranarray(Cos, (np.float32))

    # Parameters for the overlap analysis/synthesis
    N = 1024
    winsamples = 2048
    win = np.sin(np.pi/(winsamples)*(np.arange(0, winsamples) + 0.5))

    # Ovelapped analysis with 75%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N/4).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N/4).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(ms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

def spc_eval_dct_oc(mixKaraoke, vox):
    # Analysis Parameters
    N = 1024
    winsamples = 2 * N
    win = np.sin(np.pi/(winsamples)*(np.arange(0,winsamples)+0.5))

    # Acquire modulation matrices
    Cos, _ = TF.TimeFrequencyDecomposition.coreModulation(win, N)
    Cos = Cos.T

    # Create Dictionary
    CS2 = np.asfortranarray(Cos, (np.float32))

    # Ovelapped analysis with 50%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(ms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

def spc_eval_dctdst_oc(mixKaraoke, vox):
    # Analysis Parameters
    N = 1024
    winsamples = 2 * N
    win = np.sin(np.pi/(winsamples)*(np.arange(0,winsamples)+0.5))

    # Acquire modulation matrices
    Cos, Sin = TF.TimeFrequencyDecomposition.coreModulation(win, N)
    Cos = Cos.T
    Sin = Sin.T

    # Create Dictionary
    CS2 = np.asfortranarray(np.hstack((Cos, Sin)), (np.float32))

    # Ovelapped analysis with 50%
    karmix = np.asfortranarray(OLAnalysis(mixKaraoke, win, winsamples, N).T, np.float32)
    vox = np.asfortranarray(OLAnalysis(vox, win, winsamples, N).T, np.float32)

    # Acquire coefficients using orthogonal matching pursuit
    ms = np.asarray(spams.omp(karmix, CS2, 210, 1e-16).todense()).T
    vs = np.asarray(spams.omp(vox, CS2, 210, 1e-16).todense()).T

    # Compute Sparsity Criteria
    SPCMix = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(ms))
    SPCVox = TF.WDODisjointness.l1l2_sparsity_measure(np.abs(vs))

    return SPCMix, SPCVox

# Result Loading
def load_results():
    # STFT
    # Hanning
    stft_hann_a1 = np.load('WDOExperiment/stft_hann_a1.npy')
    stft_hann_a12 = np.load('WDOExperiment/stft_hann_a12.npy')

    # Bartlett
    stft_bt_a1 = np.load('WDOExperiment/stft_bt_a1.npy')
    stft_bt_a12 = np.load('WDOExperiment/stft_bt_a12.npy')

    # Nuttall-4b
    stft_nt_a1 = np.load('WDOExperiment/stft_nt_a1.npy')
    stft_nt_a12 = np.load('WDOExperiment/stft_nt_a12.npy')

    stft_ntB_a1 = np.load('WDOExperiment/stft_ntB_a1.npy')
    stft_ntB_a12 = np.load('WDOExperiment/stft_ntB_a12.npy')


    # MDCT
    mdct_a1 = np.load('WDOExperiment/mdct_a1.npy')
    mdct_a12 = np.load('WDOExperiment/mdct_a12.npy')
    mdcst_a1 = np.load('WDOExperiment/mdcst_a1.npy')
    mdcst_a12 = np.load('WDOExperiment/mdcst_a12.npy')

    # PQMF
    pqmf_cos_a1 = np.load('WDOExperiment/pqmf_cos_a1.npy')
    pqmf_cos_a2 = np.load('WDOExperiment/pqmf_cos_a12.npy')
    pqmf_compl_a1 = np.load('WDOExperiment/pqmf_compl_a1.npy')
    pqmf_compl_a2 = np.load('WDOExperiment/pqmf_compl_a12.npy')

    # Matching Pursuit OC
    mdct_oc = np.load('WDOExperiment/mdct_oc.npy')
    mdct_union_oc = np.load('WDOExperiment/mdct_union_oc.npy')
    mdcst_oc = np.load('WDOExperiment/mdcst_oc.npy')

    return stft_hann_a1,stft_hann_a12,stft_bt_a1, stft_bt_a12, stft_nt_a1,stft_nt_a12, stft_ntB_a1, stft_ntB_a12, \
    mdct_a1, mdct_a12, mdcst_a1, mdcst_a12, pqmf_cos_a1, pqmf_cos_a2, pqmf_compl_a1, pqmf_compl_a2, mdct_oc, mdcst_oc, mdct_union_oc

def load_SPCresults():
    # STFT
    # Hanning
    spc_stft_hann = np.load('WDOExperiment/spc_stft_hann_a1.npy')

    # Bartlett
    spc_stft_bt = np.load('WDOExperiment/spc_stft_bt_a1.npy')

    # Nuttall-4b
    spc_stft_nt = np.load('WDOExperiment/spc_stft_nt_a1.npy')
    spc_stft_ntB = np.load('WDOExperiment/spc_stft_ntB_a1.npy')

    # MDCT
    spc_mdct = np.load('WDOExperiment/spc_mdct_a1.npy')
    spc_mdcst = np.load('WDOExperiment/spc_mdcst_a1.npy')

    # PQMF
    spc_pqmf = np.load('WDOExperiment/spc_pqmf_cos_a1.npy')
    spc_pqmf_comp = np.load('WDOExperiment/spc_pqmf_compl_a1.npy')

    # Matching Pursuit OC
    mdct_oc = np.load('WDOExperiment/spc_mdct_oc.npy')
    mdct_union_oc = np.load('WDOExperiment/spc_mdcst_union_oc.npy')
    mdcst_oc = np.load('WDOExperiment/spc_mdcst_oc.npy')

    return spc_stft_hann, spc_stft_bt, spc_stft_nt, spc_stft_ntB, spc_mdct,\
           spc_mdcst, spc_pqmf, spc_pqmf_comp, mdct_oc, mdcst_oc, mdct_union_oc

# Main Operations
### WDO
def mainWDO(selection):
    from openpyxl import load_workbook
    MixturesPath = '/home/avdata/audio/own/dsd100/DSD100/Mixtures/'
    SourcesPath = '/home/avdata/audio/own/dsd100/DSD100/Sources/'
    savePath = '/mnt/IDMT-WORKSPACE/DATA-STORE/mis/Datasets/DSD/mono_sv_eval'
    foldersList = ['Dev', 'Test']

    xlsLoc = '/home/avdata/audio/own/dsd100/dsd100.xlsx'
    keywords = ['bass_seg.wav', 'drums_seg.wav', 'other_seg.wav', 'vocals_seg.wav']
    wb = load_workbook(filename = xlsLoc)
    sheet_ranges = wb['Sheet1']

    # Worksheet Indices
    strtIndx = 2
    endIndx = 102

    # Saving Index for storing the results
    saveIndx = 0

    # Number of sub-bands
    N = 1024

    # Initialize Results storing
    # Alpha = 1.
    stft_hann_a1 = np.zeros(endIndx - strtIndx)
    stft_bt_a1 = np.zeros(endIndx - strtIndx)
    stft_nt_a1 = np.zeros(endIndx - strtIndx)
    stft_ntB_a1 = np.zeros(endIndx - strtIndx)

    mdct_a1 = np.zeros(endIndx - strtIndx)
    mdcst_a1 = np.zeros(endIndx - strtIndx)
    pqmf_cos_a1 = np.zeros(endIndx - strtIndx)
    pqmf_compl_a1 = np.zeros(endIndx - strtIndx)

    # Alpha = 1.2
    stft_hann_a12 = np.zeros(endIndx - strtIndx)
    stft_bt_a12 = np.zeros(endIndx - strtIndx)
    stft_nt_a12 = np.zeros(endIndx - strtIndx)
    stft_ntB_a12 = np.zeros(endIndx - strtIndx)

    mdct_a12 = np.zeros(endIndx - strtIndx)
    mdcst_a12 = np.zeros(endIndx - strtIndx)
    pqmf_cos_a12 = np.zeros(endIndx - strtIndx)
    pqmf_compl_a12 = np.zeros(endIndx - strtIndx)


    # Matching Pursuit based decompositions
    mdct_oc = np.zeros(endIndx - strtIndx)
    mdct_union_oc = np.zeros(endIndx - strtIndx)
    mdcst_oc = np.zeros(endIndx - strtIndx)

    for Fileindx in range(strtIndx,endIndx):
        cell = 'A' + str(Fileindx)
        print(cell)
        strr =(sheet_ranges[cell].value)
        strrb = (sheet_ranges[cell].value)
        DevfolderName = os.path.join(MixturesPath,foldersList[0],str(strr.encode('utf-8')))
        TstfolderName = os.path.join(MixturesPath,foldersList[1],str(strrb.encode('utf-8')))

        if os.path.exists(TstfolderName):
            print('Current File belongs to Test/Validation Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[1],str(sheet_ranges[cell].value))
            folderNameMix = TstfolderName
            currFile = 'Test'

        elif os.path.exists(DevfolderName):
            print('Current File belongs to Training Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[0],str(sheet_ranges[cell].value))
            folderNameMix = DevfolderName
            currFile = 'Valid'

        print('Reading')
        mix, fs = IO.AudioIO.wavRead(os.path.join(folderNameMix, 'mixture_seg.wav'), mono = True)
        bss, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[0]), mono = True)
        drm, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[1]), mono = True)
        oth, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[2]), mono = True)
        mixKaraoke = bss + drm + oth
        vox, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[3]), mono = True)

        # Evaluation inside file loop
        if selection == 'stft':
            print('STFT')
            # Alpha = 1.
            stft_hann_a1[saveIndx] = eval_stft_hann(1., mixKaraoke, vox)
            stft_bt_a1[saveIndx] = eval_stft_bt(1., mixKaraoke, vox)
            stft_nt_a1[saveIndx] = eval_stft_nt(1., mixKaraoke, vox)
            stft_ntB_a1[saveIndx] = eval_stft_ntB(1., mixKaraoke, vox)


            # Alpha = 1.2
            stft_hann_a12[saveIndx] = eval_stft_hann(1.2, mixKaraoke, vox)
            stft_bt_a12[saveIndx] = eval_stft_bt(1.2, mixKaraoke, vox)
            stft_nt_a12[saveIndx] = eval_stft_nt(1.2, mixKaraoke, vox)
            stft_nt_a12[saveIndx] = eval_stft_ntB(1.2, mixKaraoke, vox)

            np.save('WDOExperiment/stft_hann_a1.npy', stft_hann_a1)
            np.save('WDOExperiment/stft_bt_a1.npy', stft_bt_a1)
            np.save('WDOExperiment/stft_nt_a1.npy', stft_nt_a1)
            np.save('WDOExperiment/stft_ntB_a1.npy', stft_ntB_a1)
            np.save('WDOExperiment/stft_hann_a12.npy', stft_hann_a12)
            np.save('WDOExperiment/stft_bt_a12.npy', stft_bt_a12)
            np.save('WDOExperiment/stft_nt_a12.npy', stft_nt_a12)
            np.save('WDOExperiment/stft_ntB_a12.npy', stft_ntB_a12)


        elif selection == 'mdct':
            print('MDCT')
            # Alpha = 1.
            mdct_a1[saveIndx] = eval_mdct(1., mixKaraoke, vox)
            mdcst_a1[saveIndx] = eval_mdcst_complex(1., mixKaraoke, vox)

            # Alpha = 1.2
            mdct_a12[saveIndx] = eval_mdct(1.2, mixKaraoke, vox)
            mdcst_a12[saveIndx] = eval_mdcst_complex(1.2, mixKaraoke, vox)

            np.save('WDOExperiment/mdct_a1.npy', mdct_a1)
            np.save('WDOExperiment/mdcst_a1.npy', mdcst_a1)
            np.save('WDOExperiment/mdct_a12.npy', mdct_a12)
            np.save('WDOExperiment/mdcst_a12.npy', mdcst_a12)

        elif selection == 'pqmf':
            print('PQMF')
            # Alpha = 1.
            pqmf_cos_a1[saveIndx] = eval_pqmf_cos(1., mixKaraoke, vox)

            # Alpha = 1.2
            pqmf_cos_a12[saveIndx] = eval_pqmf_cos(1.2, mixKaraoke, vox)

            np.save('WDOExperiment/pqmf_cos_a1.npy', pqmf_cos_a1)
            np.save('WDOExperiment/pqmf_cos_a12.npy', pqmf_cos_a12)


        elif selection == 'pqmf_complex':
            print('PQMF Complex')
            # Alpha = 1.
            pqmf_compl_a1[saveIndx] = eval_pqmf_complex(1., mixKaraoke, vox)

            # Alpha = 1.2
            pqmf_compl_a12[saveIndx] = eval_pqmf_complex(1.2, mixKaraoke, vox)

            np.save('WDOExperiment/pqmf_compl_a1.npy', pqmf_compl_a1)
            np.save('WDOExperiment/pqmf_compl_a12.npy', pqmf_compl_a12)

        elif selection == 'mdcst_oc':
            print('Matching Pursuit Based')
            mdct_oc[saveIndx] = eval_dct_oc(mixKaraoke, vox)
            mdct_union_oc[saveIndx] = eval_dct_union_oc(mixKaraoke, vox)
            mdcst_oc[saveIndx] = eval_dctdst_oc(mixKaraoke, vox)

            np.save('WDOExperiment/mdct_oc.npy', mdct_oc)
            np.save('WDOExperiment/mdct_union_oc.npy', mdct_union_oc)
            np.save('WDOExperiment/mdcst_oc.npy', mdcst_oc)

        else :
            assert('Unknown Selection!')

        # Update Storing Index
        saveIndx += 1

    return None

### SPC
def mainSPC(selection):
    from openpyxl import load_workbook
    MixturesPath = '/home/avdata/audio/own/dsd100/DSD100/Mixtures/'
    SourcesPath = '/home/avdata/audio/own/dsd100/DSD100/Sources/'
    savePath = '/mnt/IDMT-WORKSPACE/DATA-STORE/mis/Datasets/DSD/mono_sv_eval'
    foldersList = ['Dev', 'Test']

    xlsLoc = '/home/avdata/audio/own/dsd100/dsd100.xlsx'
    keywords = ['bass_seg.wav', 'drums_seg.wav', 'other_seg.wav', 'vocals_seg.wav']
    wb = load_workbook(filename = xlsLoc)
    sheet_ranges = wb['Sheet1']

    # Worksheet Indices
    strtIndx = 2
    endIndx = 102

    # Saving Index for storing the results
    saveIndx = 0

    # Number of sub-bands
    N = 1024

    # Initialize Results storing
    # Alpha = 1.
    stft_hann_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_bt_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_nt_a1 = np.zeros((endIndx - strtIndx, 2))
    stft_ntB_a1 = np.zeros((endIndx - strtIndx, 2))

    mdct_a1 = np.zeros((endIndx - strtIndx, 2))
    mdcst_a1 = np.zeros((endIndx - strtIndx, 2))
    pqmf_cos_a1 = np.zeros((endIndx - strtIndx, 2))
    pqmf_compl_a1 = np.zeros((endIndx - strtIndx, 2))


    mdct_oc = np.zeros((endIndx - strtIndx, 2))
    mdct_union_oc = np.zeros((endIndx - strtIndx, 2))
    mdcst_oc = np.zeros((endIndx - strtIndx, 2))

    for Fileindx in range(strtIndx,endIndx):
        cell = 'A' + str(Fileindx)
        print(cell)
        strr =(sheet_ranges[cell].value)
        strrb = (sheet_ranges[cell].value)
        DevfolderName = os.path.join(MixturesPath,foldersList[0],str(strr.encode('utf-8')))
        TstfolderName = os.path.join(MixturesPath,foldersList[1],str(strrb.encode('utf-8')))

        if os.path.exists(TstfolderName):
            print('Current File belongs to Test/Validation Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[1],str(sheet_ranges[cell].value))
            folderNameMix = TstfolderName
            currFile = 'Test'

        elif os.path.exists(DevfolderName):
            print('Current File belongs to Training Set')
            folderNameSource = os.path.join(SourcesPath,foldersList[0],str(sheet_ranges[cell].value))
            folderNameMix = DevfolderName
            currFile = 'Valid'

        print('Reading')
        mix, fs = IO.AudioIO.wavRead(os.path.join(folderNameMix, 'mixture_seg.wav'), mono = True)
        bss, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[0]), mono = True)
        drm, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[1]), mono = True)
        oth, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[2]), mono = True)
        mixKaraoke = bss + drm + oth
        vox, fs = IO.AudioIO.wavRead(os.path.join(folderNameSource, keywords[3]), mono = True)

        # Evaluation inside file loop
        if selection == 'stft':
            print('STFT')
            # Alpha = 1.
            stft_hann_a1[saveIndx, 0], stft_hann_a1[saveIndx, 1] = spc_eval_stft_hann(1., mixKaraoke, vox)
            stft_bt_a1[saveIndx, 0], stft_bt_a1[saveIndx, 1] = spc_eval_stft_bt(1., mixKaraoke, vox)
            stft_nt_a1[saveIndx, 0], stft_nt_a1[saveIndx, 1] = spc_eval_stft_nt(1., mixKaraoke, vox)
            stft_ntB_a1[saveIndx, 0], stft_ntB_a1[saveIndx, 1] = spc_eval_stft_ntB(1., mixKaraoke, vox)

            np.save('WDOExperiment/spc_stft_hann_a1.npy', stft_hann_a1)
            np.save('WDOExperiment/spc_stft_bt_a1.npy', stft_bt_a1)
            np.save('WDOExperiment/spc_stft_nt_a1.npy', stft_nt_a1)
            np.save('WDOExperiment/spc_stft_ntB_a1.npy', stft_ntB_a1)


        elif selection == 'mdct':
            print('MDCT')
            # Alpha = 1.
            mdct_a1[saveIndx, 0], mdct_a1[saveIndx, 1] = spc_eval_mdct(1., mixKaraoke, vox)
            mdcst_a1[saveIndx, 0], mdcst_a1[saveIndx, 1] = spc_eval_mdcst_complex(1., mixKaraoke, vox)

            np.save('WDOExperiment/spc_mdct_a1.npy', mdct_a1)
            np.save('WDOExperiment/spc_mdcst_a1.npy', mdcst_a1)


        elif selection == 'pqmf':
            print('PQMF')
            # Alpha = 1.
            pqmf_cos_a1[saveIndx, 0], pqmf_cos_a1[saveIndx, 1] = spc_eval_pqmf_cos(1., mixKaraoke, vox)
            print(pqmf_cos_a1[saveIndx, 0], pqmf_cos_a1[saveIndx, 1])
            np.save('WDOExperiment/spc_pqmf_cos_a1.npy', pqmf_cos_a1)


        elif selection == 'pqmf_complex':
            print('PQMF Complex')
            # Alpha = 1.
            pqmf_compl_a1[saveIndx, 0], pqmf_compl_a1[saveIndx, 1] = spc_eval_pqmf_complex(1., mixKaraoke, vox)

            np.save('WDOExperiment/spc_pqmf_compl_a1.npy', pqmf_compl_a1)

        elif selection == 'mdcst_oc':
            print('Matching Pursuit')
            mdct_oc[saveIndx, 0], mdct_oc[saveIndx, 1] = spc_eval_dct_oc(mixKaraoke, vox)
            mdct_union_oc[saveIndx, 0], mdct_union_oc[saveIndx, 1] = spc_eval_dct_union_oc(mixKaraoke, vox)
            mdcst_oc[saveIndx, 0], mdcst_oc[saveIndx, 1] = spc_eval_dctdst_oc(mixKaraoke, vox)

            np.save('WDOExperiment/spc_mdct_oc.npy', mdct_oc)
            np.save('WDOExperiment/spc_mdct_union_oc.npy', mdct_oc)
            np.save('WDOExperiment/spc_mdcst_oc.npy', mdcst_oc)

        else :
            assert('Unknown Selection!')

        # Update Storing Index
        saveIndx += 1

    return None

if __name__ == "__main__":

    # Define Operation
    # Plot acquired results
    operation = 'Results'
    # Run the disjointness experiment
    #operation = 'WDO'
    # Rim the sparsity experiment
    #operation = 'Sparsity'

    # Select multi processing feature
    multi_processing = False

    if operation == 'WDO' :
        if multi_processing == True :
            p1 = Process(target = mainWDO, args = ('stft',))
            p1.start()
            p2 = Process(target = mainWDO, args = ('mdct',))
            p2.start()
            p3 = Process(target = mainWDO, args = ('pqmf',))
            p3.start()
            p4 = Process(target = mainWDO, args = ('pqmf_complex',))
            p4.start()
            p5 = Process(target = mainWDO, args = ('mdcst_oc',))
            p5.start()

            p1.join()
            p2.join()
            p3.join()
            p4.join()
            p5.join()
        else :
            mainWDO('stft')
            mainWDO('mdct')
            mainWDO('pqmf')
            mainWDO('pqmf_complex')
            mainWDO('mdcst_oc')

    if operation == 'Sparsity' :
        if multi_processing == True :
            p1 = Process(target = mainSPC, args = ('stft',))
            p1.start()
            p2 = Process(target = mainSPC, args = ('mdct',))
            p2.start()
            p3 = Process(target = mainSPC, args = ('pqmf',))
            p3.start()
            p4 = Process(target = mainSPC, args = ('pqmf_complex',))
            p4.start()
            p5 = Process(target = mainSPC, args = ('mdcst_oc',))
            p5.start()

            p1.join()
            p2.join()
            p3.join()
            p4.join()
            p5.join()

        else :
            mainSPC('stft')
            mainSPC('mdct')
            mainSPC('pqmf')
            mainSPC('pqmf_complex')
            mainSPC('mdcst_oc')

    if operation == 'Results':
        print('Loading WDO Results')

        stft_hann_a1, stft_hann_a12, stft_bt_a1, stft_bt_a12, stft_nt_a1, stft_nt_a12, stft_ntB_a1, stft_ntB_a12, \
        mdct_a1, mdct_a12, mdcst_a1, mdcst_a12, pqmf_cos_a1, pqmf_cos_a2, pqmf_compl_a1, pqmf_compl_a2, mdct_oc,\
        mdcst_oc, mdct_union_oc = load_results()

        # Avoid two audio files
        stft_hann_a1 = np.delete(stft_hann_a1, (71, 75))
        stft_hann_a12 = np.delete(stft_hann_a12, (71, 75))
        stft_bt_a1 = np.delete(stft_bt_a1, (71, 75))
        stft_bt_a12 = np.delete(stft_bt_a12, (71, 75))
        stft_nt_a1 = np.delete(stft_nt_a1, (71, 75))
        stft_nt_a12 = np.delete(stft_nt_a12, (71, 75))
        stft_ntB_a1 = np.delete(stft_ntB_a1, (71, 75))
        stft_ntB_a12 = np.delete(stft_ntB_a12, (71, 75))
        mdct_a1 = np.delete(mdct_a1, (71, 75))
        mdct_a12 = np.delete(mdct_a12, (71, 75))
        mdcst_a1 = np.delete(mdcst_a1, (71, 75))
        mdcst_a12 = np.delete(mdcst_a12, (71, 75))
        pqmf_cos_a1 = np.delete(pqmf_cos_a1, (71, 75))
        pqmf_cos_a2 = np.delete(pqmf_cos_a2, (71, 75))
        pqmf_compl_a1 = np.delete(pqmf_compl_a1, (71, 75))
        pqmf_compl_a2 = np.delete(pqmf_compl_a2, (71, 75))
        mdct_oc = np.delete(mdct_oc, (71, 75))
        mdcst_oc = np.delete(mdcst_oc, (71, 75))
        mdct_union_oc = np.delete(mdct_union_oc, (71, 75))

        data = np.vstack((stft_hann_a1, stft_bt_a1, stft_nt_a1, stft_ntB_a1, mdct_a1, mdcst_a1,
                          pqmf_cos_a1, pqmf_compl_a1, mdct_oc, mdcst_oc, mdct_union_oc)).T

        colors = ['cyan', 'cyan', 'lightblue', 'lightblue', 'lightgreen', 'lightgreen', 'pink',
                  'pink', 'gray', 'gray', 'gray']

        labels = ['STFT Hann 50%', 'STFT Brt 50%', 'STFT Nt 75%', 'STFT Nt 50%', 'MDCT', 'MDCST',
                  'PQMF-cos', 'PQMF-complex', 'OMP DCT', 'OMP DCT/DST', 'OMP DCT-Union']

        meanpointprops = dict(marker='D', markeredgecolor='black', markerfacecolor='firebrick')

        # Configure Boxplot
        plt.figure(1)
        box = plt.boxplot(data, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('W-DO Measure')
        plt.show(block = False)

        print('Loading Sparsity Results')
        spc_stft_hann, spc_stft_bt, spc_stft_nt, spc_stft_ntB, spc_mdct, spc_mdcst, spc_pqmf,\
        spc_pqmf_comp, spc_mdct_oc, spc_mdcst_oc, spc_mdct_union_oc = load_SPCresults()

        # Acquire sparsity measures for mixture and singing voice
        spc_stft_hann_mix = np.delete(spc_stft_hann[:, 0], (71, 75))
        spc_stft_hann_sv = np.delete(spc_stft_hann[:, 1], (71,75))
        spc_stft_bt_mix = np.delete(spc_stft_bt[:, 0], (71, 75))
        spc_stft_bt_sv = np.delete(spc_stft_bt[:, 1], (71, 75))
        spc_stft_nt_mix = np.delete(spc_stft_nt[:, 0], (71, 75))
        spc_stft_nt_sv = np.delete(spc_stft_nt[:, 1], (71, 75))
        spc_stft_ntB_mix = np.delete(spc_stft_ntB[:, 0], (71, 75))
        spc_stft_ntB_sv = np.delete(spc_stft_ntB[:, 1], (71, 75))
        spc_mdct_mix = np.delete(spc_mdct[:, 0], (71, 75))
        spc_mdct_sv = np.delete(spc_mdct[:, 1], (71, 75))
        spc_mdcst_mix = np.delete(spc_mdcst[:, 0], (71, 75))
        spc_mdcst_sv = np.delete(spc_mdcst[:, 1], (71, 75))
        spc_pqmf_mix = np.delete(spc_pqmf[:, 0], (71, 75))
        spc_pqmf_sv = np.delete(spc_pqmf[:, 1], (71, 75))
        spc_pqmf_comp_mix = np.delete(spc_pqmf_comp[:, 0], (71, 75))
        spc_pqmf_comp_sv = np.delete(spc_pqmf_comp[:, 1], (71, 75))
        spc_mdct_oc_mix = np.delete(spc_mdct_oc[:, 0], (71, 75))
        spc_mdct_oc_sv = np.delete(spc_mdct_oc[:, 1], (71, 75))
        spc_mdcst_oc_mix = np.delete(spc_mdcst_oc[:, 0], (71, 75))
        spc_mdcst_oc_sv = np.delete(spc_mdcst_oc[:, 1], (71, 75))
        spc_mdct_union_oc_mix = np.delete(spc_mdct_union_oc[:, 0], (71, 75))
        spc_mdct_union_oc_sv = np.delete(spc_mdct_union_oc[:, 1], (71, 75))

        dataMIX = np.vstack((spc_stft_hann_mix, spc_stft_bt_mix, spc_stft_nt_mix, spc_stft_ntB_mix,
                             spc_mdct_mix, spc_mdcst_mix, spc_pqmf_mix, spc_pqmf_comp_mix, spc_mdct_oc_mix,
                             spc_mdcst_oc_mix, spc_mdct_union_oc_mix)).T

        dataSV = np.vstack((spc_stft_hann_sv, spc_stft_bt_sv, spc_stft_nt_sv, spc_stft_ntB_sv,
                            spc_mdct_sv, spc_mdcst_sv, spc_pqmf_sv, spc_pqmf_comp_sv, spc_mdct_oc_sv,
                            spc_mdcst_oc_sv, spc_mdct_union_oc_sv)).T

        colors = ['cyan', 'cyan', 'lightblue', 'lightblue', 'lightgreen', 'lightgreen',
                  'pink', 'pink', 'gray', 'gray', 'gray']
        labels = ['STFT Hann 50%', 'STFT Brt 50%', 'STFT Nt 75%', 'STFT Nt 50%', 'MDCT', 'MDCST', 'PQMF-cos',
                  'PQMF-complex', 'OMP DCT', 'OMP DCT/DST', 'OMP DCT-Union']

        meanpointprops = dict(marker='D', markeredgecolor='black', markerfacecolor='firebrick')

        # Configure Boxplot
        plt.figure(2)
        box2 = plt.boxplot(dataMIX, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box2['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('Sparsity Measure')
        plt.show(block = False)

        # Configure Boxplot
        plt.figure(3)
        box3 = plt.boxplot(dataSV, notch = True, patch_artist = True, labels = labels, meanprops=meanpointprops, meanline=False,
                   showmeans=True)
        for patch, color in zip(box3['boxes'], colors):
            patch.set_facecolor(color)

        plt.ylabel('Sparsity Measure')
        plt.show(block = False)